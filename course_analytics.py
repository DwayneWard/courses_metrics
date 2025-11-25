"""
Универсальный модуль для анализа ведомостей оценок курса.
Поддерживает стандартный формат Excel с ведомостями оценок.
"""

import copy
import warnings
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd

warnings.filterwarnings('ignore')


class CourseAnalytics:
    """Класс для анализа данных курса из Excel ведомости."""

    DEFAULT_ACTIVITY_TYPES: Dict[str, Dict] = {
        "project": {
            "label": "Проект",
            "patterns": ["проект", "project", "capstone"],
            "role": "primary",
            "engagement": True,
            "priority": 90,
        },
        "colloquium": {
            "label": "Коллоквиум",
            "patterns": ["коллоквиум", "colloquium"],
            "role": "primary",
            "engagement": True,
            "priority": 80,
        },
        "control": {
            "label": "Контрольная работа",
            "patterns": [
                "контрольная работа",
                "контрольные работы",
                "контрольная",
                "контрольные",
                "контрольная точка",
            ],
            "role": "primary",
            "engagement": True,
            "priority": 75,
        },
        "contest": {
            "label": "Контест / соревнование",
            "patterns": ["контест", "соревнование", "contest", "олимпиада"],
            "role": "primary",
            "engagement": True,
            "priority": 70,
        },
        "homework": {
            "label": "Домашнее задание",
            "patterns": ["домашнее задание", "домашние задания", "дз", "homework", "домашнее"],
            "role": "primary",
            "engagement": False,
            "priority": 60,
        },
        "classroom": {
            "label": "Аудиторная работа",
            "patterns": [
                "аудиторная работа",
                "аудиторные работы",
                "аудиторная",
                "аудиторные",
                "classroom",
                "auditory",
            ],
            "role": "primary",
            "engagement": False,
            "priority": 65,
        },
        "bonus": {
            "label": "Бонусная активность",
            "patterns": ["бонусная активность", "бонус", "bonus"],
            "role": "bonus",
            "engagement": False,
            "priority": 110,
        },
        "other": {
            "label": "Другое",
            "patterns": [],
            "role": "other",
            "engagement": False,
            "priority": 10,
        },
    }

    DEFAULT_CONFIG: Dict = {
        "excel_path": None,
        "stop_before": None,
        "positive_scores_only": True,
        "ignore_patterns": ["итого", "total"],
        "activity_rules": [],
        "activity_types": {},
        "primary_activity_types": None,
        "bonus_types": None,
        "engagement_activity_types": None,
        "type_order": None,
    }

    def __init__(
        self,
        excel_path: Optional[str] = None,
        stop_before_work: Optional[str] = None,
        config: Optional[Dict] = None,
    ):
        """
        Инициализация анализатора.

        Parameters:
        -----------
        excel_path : str
            Путь к Excel файлу с ведомостью
        stop_before_work : str
            Название работы, до которой учитывать данные (по умолчанию "Проект 4")
        config : dict
            Конфигурация курса (приоритетнее отдельных параметров)
        """
        self.config = self._build_config(config, excel_path, stop_before_work)
        self.excel_path = self.config["excel_path"]
        self.stop_before_work = self.config.get("stop_before")
        self.activity_rules: List[Dict] = self.config.get("activity_rules", [])
        self.activity_classifier: Dict[str, Dict] = {}
        self.bonus_types: set = set()
        self.primary_activity_types: List[str] = []
        self.engagement_activity_types: List[str] = []
        self.type_order: List[str] = []

        self._init_activity_classifier()

        self.df = None
        self.header_row = None
        self.students_df = None
        self.work_data = {}
        self.activities = {}
        self._load_data()

    @classmethod
    def _build_config(
        cls,
        user_config: Optional[Dict],
        excel_path: Optional[str],
        stop_before_work: Optional[str],
    ) -> Dict:
        config = copy.deepcopy(cls.DEFAULT_CONFIG)
        if user_config:
            config = cls._deep_update(config, user_config)
        if excel_path:
            config["excel_path"] = excel_path
        if stop_before_work is not None:
            config["stop_before"] = stop_before_work
        if config.get("excel_path") is None:
            raise ValueError("Не указан путь к Excel-файлу ведомости.")
        return config

    @staticmethod
    def _deep_update(base: Dict, updates: Dict) -> Dict:
        result = copy.deepcopy(base)
        for key, value in updates.items():
            if (
                key in result
                and isinstance(result[key], dict)
                and isinstance(value, dict)
            ):
                result[key] = CourseAnalytics._deep_update(result[key], value)
            else:
                result[key] = value
        return result

    def _init_activity_classifier(self):
        """Готовит классификатор типов активностей и производные списки."""
        types = copy.deepcopy(self.DEFAULT_ACTIVITY_TYPES)
        overrides = self.config.get("activity_types") or {}

        for type_name, override in overrides.items():
            base = types.get(
                type_name,
                {
                    "label": type_name.title(),
                    "patterns": [],
                    "role": "primary",
                    "engagement": False,
                    "priority": 50,
                },
            )
            types[type_name] = self._merge_activity_type(base, override)

        classifier = {}
        for type_name, info in types.items():
            patterns = info.get("patterns") or []
            normalized_patterns = [str(p).lower() for p in patterns if p]
            classifier[type_name] = {
                "label": info.get("label", type_name.title()),
                "patterns": normalized_patterns,
                "role": info.get("role", "other"),
                "engagement": bool(info.get("engagement", False)),
                "priority": info.get("priority", 0),
            }

        self.activity_classifier = classifier

        if self.config.get("bonus_types") is not None:
            self.bonus_types = set(self.config.get("bonus_types") or [])
        else:
            self.bonus_types = {
                name
                for name, info in classifier.items()
                if info.get("role") == "bonus"
            }

        if self.config.get("primary_activity_types") is not None:
            self.primary_activity_types = list(
                self.config.get("primary_activity_types") or []
            )
        else:
            self.primary_activity_types = [
                name
                for name, info in sorted(
                    classifier.items(),
                    key=lambda kv: kv[1].get("priority", 0),
                    reverse=True,
                )
                if info.get("role") == "primary"
            ]

        if self.config.get("engagement_activity_types") is not None:
            self.engagement_activity_types = list(
                self.config.get("engagement_activity_types") or []
            )
        else:
            engagement_types = [
                name
                for name, info in sorted(
                    classifier.items(),
                    key=lambda kv: kv[1].get("priority", 0),
                    reverse=True,
                )
                if info.get("engagement")
            ]
            self.engagement_activity_types = (
                engagement_types or list(self.primary_activity_types)
            )

        self.type_order = self._build_type_order(self.config.get("type_order"))

    @staticmethod
    def _merge_activity_type(base: Dict, override: Dict) -> Dict:
        """Сливает базовое описание типа активности с переопределением."""
        result = copy.deepcopy(base)
        for key, value in (override or {}).items():
            if key == "patterns":
                existing = result.get("patterns") or []
                if isinstance(value, list):
                    result["patterns"] = existing + value
                elif isinstance(value, str):
                    result["patterns"] = existing + [value]
            else:
                result[key] = value
        return result

    def _build_type_order(self, custom_order: Optional[List[str]]) -> List[str]:
        """Формирует приоритет типов работ."""
        if custom_order:
            return list(dict.fromkeys(custom_order))

        primary_sorted = list(self.primary_activity_types)
        bonus_sorted = sorted(
            list(self.bonus_types),
            key=lambda t: self.activity_classifier.get(t, {}).get("priority", 0),
            reverse=True,
        )
        others = sorted(
            [
                t
                for t in self.activity_classifier.keys()
                if t not in primary_sorted and t not in self.bonus_types
            ],
            key=lambda t: self.activity_classifier.get(t, {}).get("priority", 0),
            reverse=True,
        )

        order: List[str] = []
        for block in [primary_sorted, bonus_sorted, others]:
            for type_name in block:
                if type_name not in order:
                    order.append(type_name)
        if "other" not in order:
            order.append("other")
        return order

    def _fix_excel_file(self, filename: str) -> str:
        """Исправляет Excel файл, удаляя проблемные стили."""
        import zipfile
        import os
        import shutil
        from pathlib import Path
        
        fixed_filename = filename.replace('.xlsx', '_fixed.xlsx')
        
        try:
            # Создаем резервную копию
            backup = filename + '.backup'
            if not os.path.exists(backup):
                shutil.copy2(filename, backup)
            
            # Excel файл - это zip архив
            with zipfile.ZipFile(filename, 'r') as zip_ref:
                temp_dir = 'temp_excel'
                if os.path.exists(temp_dir):
                    shutil.rmtree(temp_dir)
                zip_ref.extractall(temp_dir)
                
                # Удаляем папку styles если она есть
                styles_path = os.path.join(temp_dir, 'xl', 'styles.xml')
                if os.path.exists(styles_path):
                    os.remove(styles_path)
            
            # Создаем новый архив без стилей
            with zipfile.ZipFile(fixed_filename, 'w', zipfile.ZIP_DEFLATED) as zip_out:
                for root, dirs, files in os.walk(temp_dir):
                    for file in files:
                        file_path = os.path.join(root, file)
                        arcname = os.path.relpath(file_path, temp_dir)
                        zip_out.write(file_path, arcname)
            
            # Удаляем временную директорию
            shutil.rmtree(temp_dir)
            return fixed_filename
        except Exception as e:
            print(f"Ошибка при исправлении файла: {e}")
            return filename
    
    def _load_data(self):
        """Загружает и обрабатывает данные из Excel файла."""
        try:
            self.df = pd.read_excel(self.excel_path)
        except Exception as e:
            # Пробуем исправить файл
            fixed_path = self._fix_excel_file(self.excel_path)
            try:
                self.df = pd.read_excel(fixed_path)
            except:
                raise Exception(f"Не удалось загрузить файл: {e}")
        
        # Читаем реальные названия заданий из первой строки Excel
        self._load_work_titles()
        
        # Первая строка pandas (вторая строка Excel) - заголовки типов работ
        self.header_row = self.df.iloc[0]
        # Данные студентов начинаются со второй строки pandas (третья строка Excel)
        self.students_df = self.df.iloc[1:].copy()
        self.students_df = self.students_df.reset_index(drop=True)
        # Обрабатываем структуру работ
        self._process_work_structure()
    
    def _load_work_titles(self):
        """Загружает реальные названия заданий из первой строки Excel."""
        from openpyxl import load_workbook
        
        self.work_titles = {}  # Словарь: индекс столбца -> название задания
        
        try:
            # Пробуем открыть файл напрямую
            try:
                wb = load_workbook(self.excel_path, data_only=True, read_only=True)
            except:
                # Пробуем исправленный файл
                fixed_path = self._fix_excel_file(self.excel_path)
                wb = load_workbook(fixed_path, data_only=True, read_only=True)
            
            ws = wb.active
            
            # Читаем первую строку (row 1 в Excel)
            # В pandas столбцы начинаются с индекса 0, но в Excel с 1
            # Столбцы A=1, B=2, C=3 и т.д.
            # Нас интересуют столбцы начиная с C (индекс 3 в Excel, индекс 2 в pandas)
            max_col = ws.max_column if ws.max_column else len(self.df.columns) + 1
            
            for col_idx_excel in range(3, max_col + 1):  # Начинаем с C (3)
                try:
                    cell_value = ws.cell(row=1, column=col_idx_excel).value
                    if cell_value and str(cell_value).strip():
                        # Конвертируем индекс Excel в индекс pandas (Excel 3 = pandas 2)
                        pandas_col_idx = col_idx_excel - 1
                        title = str(cell_value).strip()
                        # Убираем "Задачи для самостоятельной практики" из названия
                        title = title.replace('Задачи для самостоятельной практики', '').strip()
                        # Убираем лишние пробелы и точки в конце
                        title = title.rstrip(' .')
                        self.work_titles[pandas_col_idx] = title
                except:
                    # Если ячейка не существует, пропускаем
                    continue
            
            wb.close()
        except Exception as e:
            # Если не удалось прочитать через openpyxl, используем заголовки из pandas
            print(f"Предупреждение: не удалось прочитать названия заданий из Excel: {e}")
            self.work_titles = {}
    
    def _process_work_structure(self):
        """Обрабатывает структуру работ и собирает данные."""
        columns = self._collect_columns()
        activities = self._group_columns(columns)

        self.activities = activities
        self.work_data = {}

        for activity_name, activity_meta in activities.items():
            aggregated_scores = self._aggregate_activity_scores(
                activity_meta["columns"],
                activity_meta.get("aggregation"),
            )

            if self.config.get("positive_scores_only", True):
                scores = aggregated_scores[aggregated_scores > 0].dropna()
            else:
                scores = aggregated_scores.dropna()

            completed = len(scores)
            total = len(self.students_df)

            self.work_data[activity_name] = {
                "all_scores": aggregated_scores,
                "scores": scores,
                "completed": completed,
                "total": total,
                "column": activity_meta["columns"][0]["column_name"],
                "column_index": activity_meta["order"],
                "work_type": activity_meta["type"],
                "original_name": activity_meta.get("original_name", activity_name),
                "source_titles": activity_meta["source_titles"],
                "aggregation": activity_meta.get("aggregation"),
            }
    
    def _collect_columns(self) -> List[Dict]:
        """Собирает метаданные по столбцам с работами."""
        collected = []
        stop_index = self._find_stop_index()

        for idx in range(2, len(self.df.columns)):
            if stop_index is not None and idx >= stop_index:
                break

            header_value = self.df.iloc[0, idx]
            work_label = self._clean_string(header_value)
            if not work_label:
                continue

            raw_title = self.work_titles.get(idx, work_label)
            raw_title = self._clean_string(raw_title)

            if self._should_ignore_column(work_label, raw_title):
                continue

            column_name = self.df.columns[idx]
            all_scores = pd.to_numeric(self.students_df[column_name], errors="coerce")

            rule = self._match_activity_rule(raw_title, work_label)
            work_type = self._determine_type(raw_title, work_label, rule)

            collected.append(
                {
                    "index": idx,
                    "column_name": column_name,
                    "raw_title": raw_title,
                    "activity_label": work_label,
                    "all_scores": all_scores,
                    "rule": rule,
                    "work_type": work_type,
                }
            )
        return collected

    def _group_columns(self, columns: List[Dict]) -> Dict[str, Dict]:
        """Группирует столбцы в активности на основе правил и названий."""
        grouped: Dict[str, Dict] = {}
        
        # Проверяем, нужно ли разделять активности по неделям
        week_separation = self.config.get("week_activity_separation", False)
        week_separation_mode = self.config.get("week_separation_mode", "numbered")  # "numbered" или "mean"
        
        # Создаем словарь для отслеживания текущей недели и счетчиков активностей
        current_week = None
        week_activity_counters = {}  # {week_name: {activity_type: counter}}

        for col in columns:
            rule = col.get("rule") or {}
            raw_title = col.get("raw_title", "")
            activity_label = col.get("activity_label", "")
            
            # Определяем неделю для текущего столбца
            week_name = None
            if raw_title and "week" in str(raw_title).lower():
                # Если raw_title содержит "Week", это название недели
                week_name = raw_title
                current_week = week_name
            elif current_week:
                # Если есть текущая неделя, используем её
                week_name = current_week
            
            # Если включено разделение по неделям и есть название недели
            if week_separation and week_name:
                activity_type = activity_label if activity_label else col.get("work_type", "other")
                
                # Определяем, нужно ли нумеровать активности или использовать среднее
                if week_separation_mode == "mean":
                    # Группируем все активности одного типа в неделе вместе
                    activity_name = f"{week_name}. {activity_type}"
                    # Инициализируем счетчик для этой недели и типа активности
                    if week_name not in week_activity_counters:
                        week_activity_counters[week_name] = {}
                    if activity_type not in week_activity_counters[week_name]:
                        week_activity_counters[week_name][activity_type] = 0
                else:  # numbered
                    # Создаем уникальные имена для каждой активности
                    # Инициализируем счетчик для этой недели и типа активности
                    if week_name not in week_activity_counters:
                        week_activity_counters[week_name] = {}
                    if activity_type not in week_activity_counters[week_name]:
                        week_activity_counters[week_name][activity_type] = 0
                    
                    week_activity_counters[week_name][activity_type] += 1
                    counter = week_activity_counters[week_name][activity_type]
                    
                    if counter == 1:
                        activity_name = f"{week_name}. {activity_type}"
                    else:
                        activity_name = f"{week_name}. {activity_type} {counter}"
                
                # Если правило задает имя, используем его (но это переопределит логику)
                if rule.get("name") or rule.get("label"):
                    activity_name = rule.get("name") or rule.get("label")
            else:
                # Стандартная логика: используем имя из правила или raw_title
                activity_name = rule.get("name") or rule.get("label") or col["raw_title"]
                # Если нет правила и raw_title не содержит "Week", сбрасываем current_week
                if not rule and raw_title and "week" not in str(raw_title).lower():
                    # Проверяем, не является ли это началом новой недели
                    # (если следующий столбец будет "Week N")
                    pass  # Оставляем current_week как есть

            entry = grouped.setdefault(
                activity_name,
                {
                    "name": activity_name,
                    "columns": [],
                    "source_titles": [],
                    "order": col["index"],
                    "type": rule.get("type") or col["work_type"],
                    "aggregation": rule.get("aggregation"),
                    "original_name": rule.get("original_name", activity_name),
                },
            )

            entry["columns"].append(col)
            entry["source_titles"].append(col["raw_title"])
            entry["order"] = min(entry["order"], col["index"])

            if rule.get("type"):
                entry["type"] = rule["type"]
            elif entry["type"] is None:
                entry["type"] = col["work_type"]

            if rule.get("aggregation") and entry.get("aggregation") is None:
                entry["aggregation"] = rule["aggregation"]
            
            # Если режим "mean", устанавливаем агрегацию на среднее для активностей одного типа
            if week_separation and week_separation_mode == "mean" and week_name:
                if entry.get("aggregation") is None:
                    entry["aggregation"] = "mean"

        for activity in grouped.values():
            activity["source_titles"] = list(dict.fromkeys(activity["source_titles"]))
            activity["type"] = activity.get("type") or "other"
        return grouped

    def _aggregate_activity_scores(self, columns: List[Dict], aggregation: Optional[str]) -> pd.Series:
        """Агрегирует оценки для активности."""
        if not columns:
            return pd.Series(index=self.students_df.index, dtype=float)

        if len(columns) == 1 or not aggregation or aggregation == "single":
            return columns[0]["all_scores"]

        scores_df = pd.concat(
            [col["all_scores"].rename(f"column_{i}") for i, col in enumerate(columns)],
            axis=1,
        )

        aggregation = str(aggregation).lower()
        if aggregation == "sum":
            return scores_df.sum(axis=1, min_count=1)
        if aggregation == "mean":
            return scores_df.mean(axis=1)
        if aggregation == "max":
            return scores_df.max(axis=1, skipna=True)
        if aggregation == "min":
            return scores_df.min(axis=1, skipna=True)
        if aggregation == "last":
            return scores_df.iloc[:, -1]

        # По умолчанию оставляем первый столбец
        return columns[0]["all_scores"]

    def _find_stop_index(self) -> Optional[int]:
        """Находит индекс столбца, до которого учитываются работы."""
        if not self.stop_before_work:
            return None

        for idx in range(2, len(self.df.columns)):
            work_name = str(self.header_row.iloc[idx])
            if work_name and self.stop_before_work in work_name:
                return idx
        return None

    def _should_ignore_column(self, work_label: str, raw_title: Optional[str]) -> bool:
        """Определяет, стоит ли исключить столбец из анализа."""
        if not work_label or work_label.lower() == "nan":
            return True
        lower_label = work_label.lower()
        if "unnamed" in lower_label:
            return True

        ignore_patterns = [p.lower() for p in self.config.get("ignore_patterns", [])]
        if any(pat in lower_label for pat in ignore_patterns if pat):
            return True

        if raw_title:
            lower_raw = raw_title.lower()
            if any(pat in lower_raw for pat in ignore_patterns if pat):
                return True
        return False

    def _determine_type(
        self, raw_title: str, activity_label: str, rule: Optional[Dict] = None
    ) -> str:
        """Определяет тип работы."""
        if rule and rule.get("type"):
            return rule["type"]

        text = f"{raw_title} {activity_label}".lower()
        matched: List[Tuple[int, str]] = []

        for type_name, info in self.activity_classifier.items():
            patterns = info.get("patterns") or []
            if any(pattern in text for pattern in patterns):
                matched.append((info.get("priority", 0), type_name))

        if matched:
            matched.sort(key=lambda item: item[0], reverse=True)
            return matched[0][1]

        return "other"

    def _match_activity_rule(
        self, raw_title: str, activity_label: str
    ) -> Optional[Dict]:
        """Находит правило, подходящее под задание."""
        texts = [raw_title.lower(), activity_label.lower()]

        for rule in self.activity_rules:
            patterns = (
                rule.get("match")
                or rule.get("matches")
                or rule.get("includes")
                or rule.get("patterns")
            )
            if not patterns:
                continue
            if self._text_matches(patterns, texts):
                return rule
        return None

    @staticmethod
    def _text_matches(patterns, texts: List[str]) -> bool:
        """Проверяет наличие любого паттерна в переданных текстах."""
        if isinstance(patterns, str):
            patterns = [patterns]

        normalized_patterns = [p.lower() for p in patterns if p]

        for text in texts:
            for pattern in normalized_patterns:
                if pattern in text:
                    return True
        return False

    @staticmethod
    def _clean_string(value: Optional[str]) -> Optional[str]:
        if value is None:
            return None
        cleaned = str(value).strip()
        return cleaned if cleaned and cleaned.lower() != "nan" else None

    # ==================== ВСПОМОГАТЕЛЬНЫЕ ОТЧЕТЫ ====================

    def describe_sources(self) -> pd.DataFrame:
        """
        Возвращает таблицу с информацией о том, какие столбцы вошли в каждую активность.

        Returns
        -------
        pd.DataFrame
        """
        rows = []
        for activity_name, meta in self.activities.items():
            rows.append(
                {
                    "Активность": activity_name,
                    "Тип": meta.get("type"),
                    "Количество столбцов": len(meta.get("columns", [])),
                    "Агрегация": meta.get("aggregation") or "single",
                    "Источники": ", ".join(meta.get("source_titles", [])),
                }
            )
        return pd.DataFrame(rows)
    
    # ==================== МЕТРИКА 1: ЗАВЕРШЕННОСТЬ И АКТИВНОСТЬ ====================
    
    def get_completion_rates(self) -> pd.DataFrame:
        """
        Метрика 1.1: Процент завершения по каждому типу работы.
        Группирует работы по основным типам (проекты, коллоквиум).
        Бонусные активности исключены из основных типов оценивания.
        
        Returns:
        --------
        pd.DataFrame с колонками: Тип работы, Выполнили, Всего, Процент
        """
        # Группируем по типам работ (исключая бонусные активности)
        type_stats = {}
        
        for work_name, data in self.work_data.items():
            work_type = data['work_type']
            
            # Пропускаем бонусные активности
            if work_type in self.bonus_types:
                continue
            
            original_name = data.get('original_name', work_name)
            
            # Для проектов и коллоквиума используем оригинальное название
            display_name = original_name
            
            if display_name not in type_stats:
                type_stats[display_name] = {
                    'completed': 0,
                    'total': data['total'],
                    'work_type': work_type
                }
            
            type_stats[display_name]['completed'] += data['completed']
        
        results = []
        for work_name, stats in type_stats.items():
            results.append({
                'Работа': work_name,
                'Выполнили': stats['completed'],
                'Всего': stats['total'],
                'Процент': stats['completed'] / stats['total'] * 100,
                'Тип работы': stats['work_type']
            })
        
        df = pd.DataFrame(results)
        df = df.sort_values('Процент', ascending=False)
        return df
    
    def get_completion_rates_by_individual_works(self) -> pd.DataFrame:
        """
        Метрика 1.1 (расширенная): Процент завершения по каждой отдельной работе.
        Показывает каждую работу отдельно с реальными названиями из ведомости.
        
        Returns:
        --------
        pd.DataFrame с колонками: Работа, Выполнили, Всего, Процент, Тип работы
        """
        results = []
        for work_name, data in self.work_data.items():
            results.append({
                'Работа': work_name,
                'Выполнили': data['completed'],
                'Всего': data['total'],
                'Процент': data['completed'] / data['total'] * 100,
                'Тип работы': data['work_type'],
                'column_index': data.get('column_index', self._get_column_index(data['column']))
            })
        
        df = pd.DataFrame(results)
        if len(df) == 0:
            return df
        
        df['type_order'] = df['Тип работы'].apply(self._type_order_index)
        df = df.sort_values(['type_order', 'column_index', 'Процент'], ascending=[True, True, False])
        df = df.drop(['type_order', 'column_index'], axis=1)
        return df

    def get_engagement_trajectory(self) -> pd.DataFrame:
        """
        Метрика 1.2: Траектория вовлеченности студентов по основным работам.
        Показывает динамику вовлеченности по основным активностям в порядке их выполнения.
        
        Returns:
        --------
        pd.DataFrame с колонками: Порядок, Работа, Процент выполнения, Выполнили, Всего
        """
        main_types = self.engagement_activity_types or []
        if not main_types:
            return pd.DataFrame()

        main_works = []
        for work_name, data in self.work_data.items():
            work_type = data['work_type']
            if work_type not in main_types:
                continue
            main_works.append({
                'Работа': work_name,
                'Процент': data['completed'] / data['total'] * 100,
                'Выполнили': data['completed'],
                'Всего': data['total'],
                'work_type': work_type,
                'column_index': data.get('column_index', self._get_column_index(data['column']))
            })
        
        if not main_works:
            return pd.DataFrame()

        df = pd.DataFrame(main_works)
        df['type_order'] = df['work_type'].apply(
            lambda t: main_types.index(t) if t in main_types else len(main_types)
        )
        df = df.sort_values(['type_order', 'column_index'])
        df['Порядок'] = range(1, len(df) + 1)
        df = df.drop(['work_type', 'type_order', 'column_index'], axis=1)
        return df

    def get_bonus_engagement_trajectory(self) -> pd.DataFrame:
        """
        Метрика 1.2 (бонусная): Траектория вовлеченности по бонусным активностям.
        Показывает динамику выполнения бонусных активностей в порядке их появления в курсе.
        
        Returns:
        --------
        pd.DataFrame с колонками: Порядок, Работа, Процент выполнения, Выполнили, Всего
        """
        bonus_types = self.bonus_types or {"bonus"}
        bonus_works = []
        for work_name, data in self.work_data.items():
            if data['work_type'] in bonus_types:
                bonus_works.append({
                    'Работа': work_name,
                    'Процент': data['completed'] / data['total'] * 100,
                    'Выполнили': data['completed'],
                    'Всего': data['total'],
                    'column_index': self._get_column_index(data['column'])
                })
        
        bonus_works.sort(key=lambda x: x['column_index'])
        
        df = pd.DataFrame(bonus_works)
        if len(df) > 0:
            df['Порядок'] = range(1, len(df) + 1)
            df = df.drop('column_index', axis=1)
        return df
    
    def _get_column_index(self, column_name) -> int:
        """Получает индекс столбца для сортировки."""
        try:
            # Пытаемся найти индекс столбца в исходном DataFrame
            if hasattr(self, 'df') and column_name in self.df.columns:
                return list(self.df.columns).index(column_name)
        except:
            pass
        return 999  # Если не найдено, ставим в конец

    def _type_order_index(self, work_type: Optional[str]) -> int:
        """Возвращает индекс порядка для типа работы."""
        if work_type is None:
            return len(self.type_order)
        work_type = str(work_type)
        if work_type in self.type_order:
            return self.type_order.index(work_type)
        return len(self.type_order)

    def _extract_project_number(self, work_name: str) -> int:
        """Извлекает номер проекта из названия."""
        import re
        match = re.search(r'(\d+)', work_name)
        return int(match.group(1)) if match else 999
    
    def extract_project_number(self, work_name: str) -> int:
        """Публичный метод для извлечения номера проекта из названия."""
        return self._extract_project_number(work_name)
    
    def get_dropout_point(self) -> Dict:
        """
        Метрика 1.2 (дополнительно): Выявление точки отсева.
        
        Returns:
        --------
        dict с информацией о точке отсева
        """
        trajectory = self.get_engagement_trajectory()
        if len(trajectory) < 2:
            return {'point_found': False}
        
        # Находим максимальное падение
        trajectory['Падение'] = trajectory['Процент'].diff()
        max_drop = trajectory['Падение'].min()
        max_drop_idx = trajectory['Падение'].idxmin()
        
        return {
            'point_found': True,
            'work': trajectory.loc[max_drop_idx, 'Работа'],
            'drop_percentage': abs(max_drop),
            'before': trajectory.loc[max_drop_idx - 1, 'Процент'] if max_drop_idx > 0 else None,
            'after': trajectory.loc[max_drop_idx, 'Процент']
        }
    
    # ==================== МЕТРИКА 2: УСПЕВАЕМОСТЬ ====================
    
    def get_performance_distribution(self) -> pd.DataFrame:
        """
        Метрика 2.1: Распределение оценок по типам работ.
        
        Returns:
        --------
        pd.DataFrame с статистикой по каждому типу работы
        """
        results = []
        for work_name, data in self.work_data.items():
            scores = data['scores']
            if len(scores) > 0:
                results.append({
                    'Работа': work_name,
                    'Тип': data['work_type'],
                    'Средний балл': np.mean(scores),
                    'Медиана': np.median(scores),
                    'Стандартное отклонение': np.std(scores),
                    'Минимум': np.min(scores),
                    'Максимум': np.max(scores),
                    'Количество оценок': len(scores)
                })
        
        df = pd.DataFrame(results)
        return df
    
    def get_student_progress(self) -> pd.DataFrame:
        """
        Метрика 2.2: Прогресс студентов (динамика по основным активностям).

        Returns
        -------
        pd.DataFrame
        """
        primary_types = self.primary_activity_types or []
        if not primary_types:
            return pd.DataFrame()

        activities = []
        for work_name, data in self.work_data.items():
            work_type = data['work_type']
            if work_type not in primary_types:
                continue
            activities.append({
                'name': work_name,
                'type': work_type,
                'scores': data['all_scores'],
                'column_index': data.get('column_index', self._get_column_index(data['column'])),
            })

        if len(activities) < 2:
            return pd.DataFrame()

        activities.sort(
            key=lambda item: (
                self._type_order_index(item['type']),
                item['column_index'],
            )
        )

        activity_names = [item['name'] for item in activities]
        progress_data = []

        for idx in range(len(self.students_df)):
            student_row = {}
            ordered_scores = []

            for activity in activities:
                scores = activity['scores']
                value = scores.iloc[idx] if idx < len(scores) else np.nan
                student_row[activity['name']] = value
                ordered_scores.append(value)

            valid_scores = [val for val in ordered_scores if not pd.isna(val) and val > 0]

            if len(valid_scores) >= 2:
                progress = valid_scores[-1] - valid_scores[0]
                trend = (
                    'улучшение'
                    if progress > 0
                    else 'ухудшение'
                    if progress < 0
                    else 'стабильно'
                )
            else:
                progress = np.nan
                trend = 'недостаточно данных'

            student_row['Прогресс'] = progress
            student_row['Тренд'] = trend
            student_row['Студент'] = idx
            progress_data.append(student_row)

        return pd.DataFrame(progress_data, columns=activity_names + ['Прогресс', 'Тренд', 'Студент'])
    
    def get_progress_summary(self) -> Dict:
        """
        Сводная статистика по прогрессу студентов.
        
        Returns:
        --------
        dict со статистикой прогресса
        """
        progress_df = self.get_student_progress()
        if 'Прогресс' not in progress_df.columns:
            return {}
        
        progress = progress_df['Прогресс'].dropna()
        return {
            'Улучшилось': (progress > 0).sum(),
            'Ухудшилось': (progress < 0).sum(),
            'Стабильно': (progress == 0).sum(),
            'Средний прогресс': progress.mean(),
            'Медиана прогресса': progress.median()
        }
    
    # ==================== МЕТРИКА 3: КАЧЕСТВО ОБУЧЕНИЯ ====================
    
    def get_correlations(self) -> pd.DataFrame:
        """
        Метрика 3.1: Корреляция между основными активностями (траектория вовлеченности).
        
        Returns:
        --------
        pd.DataFrame с корреляционной матрицей
        """
        engagement_types = set(self.engagement_activity_types or [])
        if not engagement_types:
            return pd.DataFrame()

        work_scores = {}
        for work_name, data in self.work_data.items():
            if data['work_type'] in engagement_types:
                work_scores[work_name] = data['all_scores']

        if len(work_scores) < 2:
            return pd.DataFrame()

        corr_df = pd.DataFrame(work_scores)
        corr_matrix = corr_df.corr()

        return corr_matrix
    
    def get_correlations_with_projects(self) -> pd.DataFrame:
        """
        Корреляции primary-активностей (не входящих в траекторию) и bonus-активностей
        с primary-активностями траектории вовлеченности.

        Returns
        -------
        pd.DataFrame
            Строки — источники (primary вне engagement + bonus), столбцы — engagement (primary)
        """
        engagement_types = set(self.engagement_activity_types or [])
        primary_types = set(self.primary_activity_types or [])
        bonus_types_set = self.bonus_types or {'bonus'}

        if not engagement_types or not primary_types:
            return pd.DataFrame()

        engagement_scores = {}
        source_scores = {}

        for work_name, data in self.work_data.items():
            work_type = data['work_type']
            if work_type in engagement_types and work_type in primary_types:
                engagement_scores[work_name] = data['all_scores']
            elif work_type in primary_types:
                source_scores[work_name] = data['all_scores']
            elif work_type in bonus_types_set:
                source_scores[work_name] = data['all_scores']

        if len(engagement_scores) == 0 or len(source_scores) == 0:
            return pd.DataFrame()

        combined = {**engagement_scores, **source_scores}
        corr_df = pd.DataFrame(combined)

        correlations = {}
        for source_work in source_scores.keys():
            correlations[source_work] = {}
            for engagement_work in engagement_scores.keys():
                corr_value = corr_df[source_work].corr(corr_df[engagement_work])
                correlations[source_work][engagement_work] = corr_value

        result_df = pd.DataFrame(correlations).T
        return result_df[engagement_scores.keys()]
    
    def get_consistency(self) -> pd.DataFrame:
        """
        Метрика 3.2: Консистентность результатов студентов по основным активностям (primary).

        Returns:
        --------
        pd.DataFrame с консистентностью для каждого студента
        """
        primary_types = set(self.primary_activity_types or [])
        if not primary_types:
            return pd.DataFrame()

        primary_scores = {
            work_name: data['all_scores']
            for work_name, data in self.work_data.items()
            if data['work_type'] in primary_types
        }

        if len(primary_scores) == 0:
            return pd.DataFrame()

        consistency_data = []
        for i in range(len(self.students_df)):
            student_name_col = self.students_df.columns[0]
            student_full_info = str(self.students_df.iloc[i][student_name_col])
            student_name = student_full_info.split('\n')[0].split('@')[0].strip()

            student_scores = []
            scored_activities = {}
            for activity_name, scores in primary_scores.items():
                if i < len(scores):
                    score = scores.iloc[i]
                    if not pd.isna(score) and score > 0:
                        student_scores.append(score)
                        scored_activities[activity_name] = score

            attempts_count = len(student_scores)

            if attempts_count == 0:
                mean = np.nan
                std = np.nan
                cv = np.nan
                category = 'нет активностей'
            elif attempts_count == 1:
                mean = student_scores[0]
                std = np.nan
                cv = np.nan
                category = 'единичная активность'
            else:
                std = np.std(student_scores)
                mean = np.mean(student_scores)
                cv = std / mean if mean > 0 else np.nan

                if cv < 0.1 and mean >= 5:
                    category = 'стабильно успешный'
                elif cv < 0.1 and mean < 5:
                    category = 'стабильно неуспешный'
                elif cv > 0.3:
                    category = 'нестабильный'
                elif mean < 5:
                    category = 'стабильно неуспешный'
                else:
                    category = 'умеренно стабильный'

            consistency_entry = {
                'Студент': i,
                'ФИО': student_name,
                'Средний балл': mean if not pd.isna(mean) else np.nan,
                'Стандартное отклонение': std,
                'Коэффициент вариации': cv,
                'Категория': category,
                'Количество активностей': attempts_count,
            }
            for activity_name, score in scored_activities.items():
                consistency_entry[f'Активность: {activity_name}'] = score

            consistency_data.append(consistency_entry)

        return pd.DataFrame(consistency_data)
    
    # ==================== МЕТРИКА 6: СРАВНИТЕЛЬНЫЕ МЕТРИКИ ====================
    
    def get_difficulty_comparison(self) -> pd.DataFrame:
        """
        Метрика 6.1: Сравнение сложности основных (primary) активностей.

        Returns:
        --------
        pd.DataFrame с индексом сложности для каждой активности
        """
        primary_types = set(self.primary_activity_types or [])
        if not primary_types:
            return pd.DataFrame()

        results = []
        for work_name, data in self.work_data.items():
            work_type = data['work_type']
            if work_type not in primary_types:
                continue

            scores = data['scores']
            completion_rate = data['completed'] / data['total'] * 100 if data['total'] else 0

            if len(scores) > 0:
                mean_score = float(np.mean(scores)) if len(scores) > 0 else np.nan
                std_score = float(np.std(scores)) if len(scores) > 0 else np.nan

                results.append({
                    'Активность': work_name,
                    'Тип': work_type,
                    'Тип (название)': self.activity_classifier.get(work_type, {}).get('label', work_type),
                    'Процент выполнения': completion_rate,
                    'Средний балл': mean_score,
                    'Стандартное отклонение': std_score,
                    'Выполнили': data['completed'],
                    'Всего': data['total']
                })

        df = pd.DataFrame(results)
        if len(df) > 0:
            def _normalize(series: pd.Series, reverse: bool = False) -> pd.Series:
                if series.max() == series.min():
                    return pd.Series([0.5] * len(series), index=series.index)
                if reverse:
                    return (series.max() - series) / (series.max() - series.min())
                return (series - series.min()) / (series.max() - series.min())

            df['Компонент (выполнение)'] = _normalize(df['Процент выполнения'], reverse=True)
            df['Компонент (качество)'] = _normalize(df['Средний балл'], reverse=True)
            df['Компонент (разброс)'] = _normalize(df['Стандартное отклонение'], reverse=False)

            weights = {
                'Компонент (выполнение)': 0.4,
                'Компонент (качество)': 0.4,
                'Компонент (разброс)': 0.2
            }

            df['Индекс сложности'] = (
                df['Компонент (выполнение)'] * weights['Компонент (выполнение)'] +
                df['Компонент (качество)'] * weights['Компонент (качество)'] +
                df['Компонент (разброс)'] * weights['Компонент (разброс)']
            ) / sum(weights.values())

            df = df.sort_values('Индекс сложности', ascending=False)
        return df
    
    def get_variability_analysis(self) -> pd.DataFrame:
        """
        Метрика 6.2: Разброс оценок (вариативность).
        
        Returns:
        --------
        pd.DataFrame с анализом вариативности по каждому типу работы
        """
        results = []
        for work_name, data in self.work_data.items():
            scores = data['scores']
            if len(scores) > 0:
                mean_score = np.mean(scores)
                std_score = np.std(scores)
                cv = std_score / mean_score if mean_score > 0 else np.nan
                
                results.append({
                    'Работа': work_name,
                    'Тип': data['work_type'],
                    'Среднее': mean_score,
                    'Стандартное отклонение': std_score,
                    'Коэффициент вариации': cv,
                    'Интерпретация': self._interpret_variability(cv, mean_score)
                })
        
        df = pd.DataFrame(results)
        return df
    
    def _interpret_variability(self, cv: float, mean: float) -> str:
        """Интерпретирует коэффициент вариации."""
        if pd.isna(cv):
            return 'недостаточно данных'
        if cv < 0.1:
            return 'низкий разброс (задание слишком простое/сложное для всех)'
        elif cv < 0.3:
            return 'умеренный разброс (хорошая дифференциация)'
        else:
            return 'высокий разброс (большая дифференциация результатов)'
    
    # ==================== СВОДНЫЕ МЕТРИКИ ====================
    
    def get_summary(self) -> Dict:
        """Возвращает сводную информацию по курсу."""
        completion = self.get_completion_rates()
        performance = self.get_performance_distribution()
        
        return {
            'Количество студентов': len(self.students_df),
            'Количество работ': len(self.work_data),
            # 'Средний процент завершения': completion['Процент'].mean(),
            # 'Средний балл по всем работам': performance['Средний балл'].mean() if len(performance) > 0 else 0,
            'Работ до учета': self.stop_before_work
        }
    
    def get_all_metrics(self) -> Dict:
        """Возвращает все рассчитанные метрики."""
        return {
            'summary': self.get_summary(),
            'completion_rates': self.get_completion_rates(),
            'engagement_trajectory': self.get_engagement_trajectory(),
            'dropout_point': self.get_dropout_point(),
            'performance_distribution': self.get_performance_distribution(),
            'student_progress': self.get_student_progress(),
            'progress_summary': self.get_progress_summary(),
            'correlations': self.get_correlations(),
            'consistency': self.get_consistency(),
            'difficulty_comparison': self.get_difficulty_comparison(),
            'variability_analysis': self.get_variability_analysis()
        }

